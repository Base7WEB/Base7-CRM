import csv
import io

from openpyxl import load_workbook

from whatsapp import normalizar_telefone

EXTENSOES_PERMITIDAS = {"csv", "xlsx", "txt"}
TAMANHO_MAXIMO_BYTES = 5 * 1024 * 1024  # 5 MB

_APELIDOS = {
    "empresa":     ["empresa", "nome", "company", "razao social", "razão social"],
    "responsavel": ["responsavel", "responsável", "contato", "nome do contato"],
    "telefone":    ["telefone", "whatsapp", "celular", "phone", "fone", "tel"],
    "cidade":      ["cidade", "city", "municipio", "município"],
    "nicho":       ["nicho", "segmento", "categoria", "ramo"],
}


def extensao_valida(nome_arquivo):
    ext = nome_arquivo.rsplit(".", 1)[-1].lower() if "." in nome_arquivo else ""
    return ext in EXTENSOES_PERMITIDAS, ext


def processar_arquivo(nome_arquivo, conteudo):
    _, ext = extensao_valida(nome_arquivo)
    if ext == "csv":
        return _processar_csv(conteudo)
    if ext == "xlsx":
        return _processar_xlsx(conteudo)
    if ext == "txt":
        return _processar_txt(conteudo)
    raise ValueError("Formato de arquivo não suportado")


def _processar_csv(conteudo):
    texto = conteudo.decode("utf-8-sig", errors="replace")
    amostra = texto[:2048]
    try:
        dialect = csv.Sniffer().sniff(amostra, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    linhas_brutas = list(csv.reader(io.StringIO(texto), dialect))
    linhas_brutas = [linha for linha in linhas_brutas if any(c.strip() for c in linha)]
    if not linhas_brutas:
        return {"colunas": [], "linhas": [], "mapeamento_sugerido": {}}

    colunas = [c.strip() for c in linhas_brutas[0]]
    linhas  = linhas_brutas[1:]
    return {"colunas": colunas, "linhas": linhas, "mapeamento_sugerido": _sugerir_mapeamento(colunas)}


def _processar_xlsx(conteudo):
    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    linhas_brutas = [list(linha) for linha in ws.iter_rows(values_only=True)]
    linhas_brutas = [linha for linha in linhas_brutas if any(c not in (None, "") for c in linha)]
    if not linhas_brutas:
        return {"colunas": [], "linhas": [], "mapeamento_sugerido": {}}

    colunas = [str(c).strip() if c is not None else "" for c in linhas_brutas[0]]
    linhas  = [["" if c is None else str(c) for c in linha] for linha in linhas_brutas[1:]]
    return {"colunas": colunas, "linhas": linhas, "mapeamento_sugerido": _sugerir_mapeamento(colunas)}


def _processar_txt(conteudo):
    """Aceita uma lista de telefones (um por linha) ou 'Nome | Telefone'."""
    texto  = conteudo.decode("utf-8", errors="replace")
    linhas = []
    for linha_bruta in texto.splitlines():
        linha_bruta = linha_bruta.strip()
        if not linha_bruta:
            continue
        if "|" in linha_bruta:
            partes = [p.strip() for p in linha_bruta.split("|", 1)]
            empresa, telefone = partes[0], (partes[1] if len(partes) > 1 else "")
        else:
            empresa, telefone = "", linha_bruta
        linhas.append([empresa, telefone])

    return {"colunas": ["empresa", "telefone"], "linhas": linhas,
            "mapeamento_sugerido": {"empresa": 0, "telefone": 1}}


def _sugerir_mapeamento(colunas):
    mapeamento = {}
    colunas_lower = [c.lower().strip() for c in colunas]
    for campo, apelidos in _APELIDOS.items():
        for idx, col in enumerate(colunas_lower):
            if col in apelidos:
                mapeamento[campo] = idx
                break
    return mapeamento


def montar_leads(linhas, mapeamento):
    """Aplica o mapeamento de colunas {campo: indice_coluna} sobre as linhas cruas."""
    leads = []
    for linha in linhas:
        lead = {}
        for campo, idx in mapeamento.items():
            if idx is None or idx == "":
                continue
            idx = int(idx)
            if idx < len(linha):
                lead[campo] = str(linha[idx] or "").strip()
        lead["telefone_normalizado"] = normalizar_telefone(lead.get("telefone", ""))
        if lead.get("empresa") or lead.get("telefone"):
            leads.append(lead)
    return leads
